"""
Excel Generator - Create Excel files with extracted voter data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
from typing import List, Dict


def generate_excel(data: List[Dict], output_path: str) -> bool:
    """
    Generate Excel file from extracted data
    
    Args:
        data: Array of extracted voter data
        output_path: Path to save Excel file
    
    Returns:
        True if successful, raises exception on error
    """
    try:
        # Create workbook and worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Voter Data'
        
        # Determine all image columns
        include_main_image = any(record.get('image_base64') for record in data)
        additional_image_keys = set()
        for record in data:
            for key in record.keys():
                if key.endswith('_image'):
                    additional_image_keys.add(key)
        
        sorted_image_keys = sorted(list(additional_image_keys))
        
        # Define base headers
        headers = [
            'Serial No', 
            'EPIC No', 
            'Name', 'Name (English)', 'Name (Kannada)',
            'Relation Type', 
            'Relative Name', 'Relative Name (English)', 'Relative Name (Kannada)',
            'House No', 
            'Gender', 
            'Age', 'Assembly No',
            'Booth Center', 'Booth Center (English)', 'Booth Center (Kannada)',
            'Booth Address', 'Booth Address (English)', 'Booth Address (Kannada)',
            'Prabhag', 'Booth No'
        ]
        
        if include_main_image:
            headers.append('Base64 Image String')
        
        for key in sorted_image_keys:
            # Format header name
            header_name = key.replace('_', ' ').title()
            headers.append(header_name)
        
        # Apply styling to headers
        header_font = Font(bold=True, size=12, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
  
        worksheet.append(headers) 
  
        # Style header row and set column widths
        for col_num in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # Set default widths for standard columns, then auto for images
            col_letter = get_column_letter(col_num)
            if col_num <= 21: # Increased from 19 to 21 because of 2 new columns
                standard_widths = [10, 15, 25, 25, 25, 15, 25, 25, 25, 10, 8, 8, 12, 30, 30, 30, 30, 30, 30, 15, 15]
                worksheet.column_dimensions[col_letter].width = standard_widths[col_num-1]
            else:
                worksheet.column_dimensions[col_letter].width = 25 # Image B64 columns
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sort data by Page, then Row, then Column (Perfect Spatial Row-wise Order)
        def get_spatial_key(rec):
            # Page, Row, Column (1-indexed based on dict)
            return (
                rec.get('page', 0), 
                rec.get('row', 0), 
                rec.get('column', 0)
            )

        data.sort(key=get_spatial_key)

        # Add data rows
        for index, record in enumerate(data):
            row_num = index + 2
            
            # Prepare row values in order
            row_values = [
                record.get('serialNo') or record.get('Serial No') or record.get('serial_no') or '',
                record.get('voterID', ''),
                record.get('name', ''),
                record.get('nameEnglish', ''),
                record.get('nameKannada', ''),
                record.get('relationType', ''),
                record.get('relativeName', ''),
                record.get('relativeNameEnglish', ''),
                record.get('relativeNameKannada', ''),
                record.get('houseNo', ''),
                record.get('gender', ''),
                record.get('age', ''),
                record.get('assemblyNo', ''),
                record.get('boothCenter', ''),
                record.get('boothCenterEnglish', ''),
                record.get('boothCenterKannada', ''),
                record.get('boothAddress', ''),
                record.get('boothAddressEnglish', ''),
                record.get('boothAddressKannada', ''),
                record.get('prabhag', ''),
                record.get('boothNo', '')
            ]

            if include_main_image:
                row_values.append(record.get('image_base64', ''))
            
            for key in sorted_image_keys:
                row_values.append(record.get(key, ''))
            
            # Write row
            worksheet.append(row_values)
            
            # Style the row
            for col_num in range(1, len(headers) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                
                cell.alignment = Alignment(vertical='top')

                # Alternate row colors
                if index % 2 == 0:
                    light_gray = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
                    cell.fill = light_gray
        
        # Save workbook
        workbook.save(output_path)
        print(f"Excel file generated: {output_path}")
        
        return True
    
    except Exception as e:
        print(f"Excel generation error: {str(e)}")
        raise

